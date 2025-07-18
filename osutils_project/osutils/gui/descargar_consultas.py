import customtkinter as ctk
from tkinter import filedialog
from osutils.gui.conexion import SelectorConexionOracle
import os
import threading
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

class DescargarConsultasFrame(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master)
        content_frame = ctk.CTkFrame(self, fg_color="transparent")
        content_frame.pack(pady=10, padx=30, fill="both", expand=True)
        self.selector = SelectorConexionOracle(content_frame)
        self.selector.pack(pady=10, anchor="w")
        ctk.CTkLabel(content_frame, text="Archivo de consultas (.sql):").pack(pady=5, anchor="w")
        sql_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
        sql_frame.pack(pady=5, anchor="w")
        self.sql_entry = ctk.CTkEntry(sql_frame, width=320)
        self.sql_entry.pack(side="left", padx=(0, 5))
        self.btn_explorar_sql = ctk.CTkButton(
            sql_frame,
            text="📂",
            width=40,
            height=28,
            command=self.seleccionar_archivo_sql
        )
        self.btn_explorar_sql.pack(side="left")
        ctk.CTkLabel(content_frame, text="Ruta de archivo Excel de salida:").pack(pady=5, anchor="w")
        out_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
        out_frame.pack(pady=5, anchor="w")
        self.out_entry = ctk.CTkEntry(out_frame, width=320)
        self.out_entry.pack(side="left", padx=(0, 5))
        self.btn_explorar_out = ctk.CTkButton(
            out_frame,
            text="📂",
            width=40,
            height=28,
            command=self.seleccionar_archivo_out
        )
        self.btn_explorar_out.pack(side="left")
        self.btn_ejecutar = ctk.CTkButton(content_frame, text="Descargar Consultas", command=self.ejecutar_consultas)
        self.btn_ejecutar.pack(pady=20, anchor="w")
        self.mensaje = ctk.CTkLabel(content_frame, text="", font=("Arial", 14))
        self.mensaje.pack(pady=10, anchor="w")
        self.texto_avance = ctk.CTkTextbox(content_frame, height=200, width=600)
        self.texto_avance.pack(pady=10, anchor="w")
    def seleccionar_archivo_sql(self):
        ruta = filedialog.askopenfilename(
            title="Selecciona el archivo .sql",
            filetypes=[("SQL files", "*.sql")]
        )
        if ruta:
            self.sql_entry.delete(0, "end")
            self.sql_entry.insert(0, ruta)
    def seleccionar_archivo_out(self):
        # Sugerir nombre de archivo con fecha actual
        fecha_str = datetime.now().strftime("%d%m%Y")
        sugerido = f"resultados_{fecha_str}.xlsx"
        ruta = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=sugerido,
            title="Selecciona la ruta de salida"
        )
        if ruta:
            self.out_entry.delete(0, "end")
            self.out_entry.insert(0, ruta)
    def ejecutar_consultas(self):
        self.mensaje.configure(text="")
        self.texto_avance.delete("1.0", "end")
        sql_path = self.sql_entry.get().strip()
        out_path = self.out_entry.get().strip()
        conexion_nombre = self.selector.get_selected()
        if not sql_path or not os.path.exists(sql_path):
            self.mensaje.configure(text="❌ Debes seleccionar un archivo .sql válido.")
            return
        if not out_path:
            self.mensaje.configure(text="❌ Debes seleccionar la ruta de salida.")
            return
        if not conexion_nombre:
            self.mensaje.configure(text="❌ Debes seleccionar una conexión.")
            return
        def run_consultas():
            try:
                with open(sql_path, "r", encoding="utf-8") as f:
                    contenido = f.read()
                # Separar consultas por ; y limpiar
                consultas = [q.strip() for q in contenido.split(";") if q.strip()]
                if not consultas:
                    self.mensaje.configure(text="No se encontraron consultas en el archivo.")
                    return
                connection = self.selector.connect()
                cursor = connection.cursor()
                wb = openpyxl.Workbook()
                ws_sql = wb.create_sheet("SQL")
                ws_sql.append(["Consultas ejecutadas"])
                for idx, consulta in enumerate(consultas):
                    nombre_hoja = f"Consulta_{idx+1}"
                    ws = wb.create_sheet(nombre_hoja)
                    self.texto_avance.insert("end", f"Ejecutando consulta {idx+1}: {consulta}\n")
                    try:
                        cursor.execute(consulta)
                        columns = [col[0] for col in cursor.description]
                        ws.append(columns)
                        for row in cursor.fetchall():
                            fila = []
                            for v in row:
                                # Formatear fechas
                                if hasattr(v, 'strftime'):
                                    fila.append(v.strftime("%d/%m/%Y %H:%M:%S"))
                                else:
                                    fila.append(v)
                            ws.append(fila)
                        ws_sql.append([consulta])
                    except Exception as e:
                        ws.append([f"Error: {str(e)}"])
                        ws_sql.append([f"Error en consulta {idx+1}: {str(e)}"])
                        self.texto_avance.insert("end", f"❌ Error en consulta {idx+1}: {str(e)}\n")
                cursor.close()
                connection.close()
                # Eliminar la hoja por defecto si está vacía
                if wb.sheetnames[0] == "Sheet" and wb["Sheet"].max_row == 1:
                    wb.remove(wb["Sheet"])
                wb.save(out_path)
                self.mensaje.configure(text=f"✅ Consultas descargadas en {out_path}")
            except Exception as e:
                self.mensaje.configure(text=f"❌ Error: {str(e)}")
        threading.Thread(target=run_consultas, daemon=True).start()
